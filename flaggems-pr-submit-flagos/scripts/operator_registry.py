#!/usr/bin/env python3
"""FlagGems 算子规范名查询 + PR 链接回填工具。

数据源:
    - /workspace/规范名.xlsx          → 规范命名映射 + PR 链接回填
    - /workspace/第一批pr算子.xlsx     → 待提交算子列表 + 加速比

用法:
    # 查询单个算子的规范名和加速比
    python operator_registry.py lookup _cholesky_solve_helper

    # 回填 PR 链接
    python operator_registry.py backfill _cholesky_solve_helper https://github.com/flagos-ai/FlagGems/pull/3354

    # 列出待提交的算子（未回填 PR 链接的）
    python operator_registry.py pending [--limit 20]

    # 列出已提交的算子（已回填 PR 链接的）
    python operator_registry.py submitted
"""

import argparse
import os
import sys

import openpyxl
import pandas as pd


NORM_XLSX = os.environ.get("FLAGGEMS_NORM_XLSX", "/workspace/规范名.xlsx")
PR_XLSX = os.environ.get("FLAGGEMS_PR_XLSX", "/workspace/第一批pr算子.xlsx")


def load_norm_names():
    df = pd.read_excel(NORM_XLSX, sheet_name="Sheet1")
    df.columns = [c.replace("\r", "").replace("_x000d_", "").strip() for c in df.columns]
    df["算子名"] = df["算子名"].astype(str).str.strip()
    df["规范名"] = df["算子名（规范命名）"].astype(str).str.strip()
    if "代码路径" not in df.columns:
        df["代码路径"] = ""
    else:
        df["代码路径"] = df["代码路径"].astype(str).str.replace("_x000d_", "", regex=False).str.strip()
    df.loc[df["代码路径"] == "nan", "代码路径"] = ""
    return df


def load_pr_list():
    df = pd.read_excel(PR_XLSX, sheet_name="Sheet1")
    df["算子名称"] = df["算子名称"].astype(str).str.strip()
    speed_col = [c for c in df.columns if "加速比" in c][0]
    df["加速比"] = df[speed_col].astype(str).str.replace("_x000d_", "").str.strip()
    df["纯名"] = df["算子名称"].str.replace("aten::", "", regex=False).str.strip()
    return df


def lookup(op_name):
    """查询算子的规范名、加速比、PR状态"""
    norm_df = load_norm_names()
    pr_df = load_pr_list()

    clean_name = op_name.replace("aten::", "").strip()

    norm_row = norm_df[norm_df["算子名"] == clean_name]
    pr_row = pr_df[pr_df["纯名"] == clean_name]

    result = {"原名": clean_name}

    if len(norm_row) > 0:
        row = norm_row.iloc[0]
        result["规范名"] = row["规范名"]
        result["PR链接"] = row["代码路径"] if row["代码路径"] else "(未提交)"
        result["行号"] = int(norm_row.index[0]) + 2  # Excel 行号 (1-based + header)
    else:
        result["规范名"] = clean_name
        result["PR链接"] = "(未在规范名表中)"
        result["行号"] = None

    if len(pr_row) > 0:
        result["加速比"] = pr_row.iloc[0]["加速比"]
        result["在待提交列表"] = True
    else:
        result["加速比"] = "(未在待提交列表中)"
        result["在待提交列表"] = False

    return result


def backfill(op_name, pr_url):
    """将 PR 链接回填到规范名.xlsx 的代码路径列"""
    clean_name = op_name.replace("aten::", "").strip()

    wb = openpyxl.load_workbook(NORM_XLSX)
    ws = wb["Sheet1"]

    header = [cell.value for cell in ws[1]]
    name_col = None
    path_col = None
    for i, h in enumerate(header):
        if h and str(h).strip() == "算子名":
            name_col = i
        if h and "代码路径" in str(h).strip():
            path_col = i

    if name_col is None or path_col is None:
        print(f"错误: 未找到列 (算子名={name_col}, 代码路径={path_col})")
        return False

    found = False
    for row in ws.iter_rows(min_row=2):
        cell_val = str(row[name_col].value or "").strip()
        if cell_val == clean_name:
            row[path_col].value = pr_url
            found = True
            print(f"✓ 已回填: {clean_name} → {pr_url} (行 {row[0].row})")
            break

    if not found:
        print(f"✗ 未找到算子 '{clean_name}' in 规范名.xlsx")
        return False

    wb.save(NORM_XLSX)
    print(f"✓ 已保存 {NORM_XLSX}")
    return True


def pending(limit=None):
    """列出待提交的算子（在第一批列表中但未回填PR链接的）"""
    norm_df = load_norm_names()
    pr_df = load_pr_list()

    merged = pr_df.merge(
        norm_df[["算子名", "规范名", "代码路径"]],
        left_on="纯名",
        right_on="算子名",
        how="left",
    )

    not_submitted = merged[
        (merged["代码路径"].isna()) | (merged["代码路径"] == "")
    ].copy()
    not_submitted = not_submitted.sort_values("加速比", ascending=False)

    if limit:
        not_submitted = not_submitted.head(limit)

    print(f"待提交算子（共 {len(not_submitted)} 个）:\n")
    print(f"{'算子名':<45} {'规范名':<45} {'加速比':>10}")
    print("-" * 100)
    for _, row in not_submitted.iterrows():
        name = row["纯名"]
        norm = row.get("规范名", name)
        if pd.isna(norm):
            norm = name
        speed = row["加速比"]
        print(f"{name:<45} {norm:<45} {speed:>10}")


def submitted():
    """列出已提交的算子"""
    norm_df = load_norm_names()
    pr_df = load_pr_list()

    merged = pr_df.merge(
        norm_df[["算子名", "规范名", "代码路径"]],
        left_on="纯名",
        right_on="算子名",
        how="left",
    )

    done = merged[
        (merged["代码路径"].notna()) & (merged["代码路径"] != "")
    ].copy()

    print(f"已提交算子（共 {len(done)} 个）:\n")
    print(f"{'算子名':<40} {'PR链接':<60}")
    print("-" * 100)
    for _, row in done.iterrows():
        name = row["纯名"]
        url = row["代码路径"]
        print(f"{name:<40} {url:<60}")


def main():
    parser = argparse.ArgumentParser(description="FlagGems 算子规范名查询 + PR回填")
    sub = parser.add_subparsers(dest="command")

    p_lookup = sub.add_parser("lookup", help="查询算子信息")
    p_lookup.add_argument("operator", help="算子名（可带 aten:: 前缀）")

    p_backfill = sub.add_parser("backfill", help="回填 PR 链接")
    p_backfill.add_argument("operator", help="算子名")
    p_backfill.add_argument("pr_url", help="PR 链接")

    p_pending = sub.add_parser("pending", help="列出待提交算子")
    p_pending.add_argument("--limit", type=int, default=None, help="限制输出数量")

    sub.add_parser("submitted", help="列出已提交算子")

    args = parser.parse_args()

    if args.command == "lookup":
        result = lookup(args.operator)
        print(f"算子名:     {result['原名']}")
        print(f"规范名:     {result['规范名']}")
        print(f"加速比:     {result['加速比']}")
        print(f"PR链接:     {result['PR链接']}")
        if result["行号"]:
            print(f"规范名表行: {result['行号']}")
    elif args.command == "backfill":
        backfill(args.operator, args.pr_url)
    elif args.command == "pending":
        pending(args.limit)
    elif args.command == "submitted":
        submitted()
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
